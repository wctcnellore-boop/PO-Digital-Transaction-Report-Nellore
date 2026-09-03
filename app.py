from datetime import datetime, timedelta
import io
import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
import pandas as pd
import streamlit as st

st.set_page_config(page_title="63 Offices Daily Report Generator", layout="centered")

st.title("📊 Daily Digital Transactions Report Generator")
st.write(
    "Upload your raw daily data file below to instantly generate the clean,"
    " formatted Excel report for all 63 offices."
)

uploaded_file = st.file_uploader(
    "Upload DailyData Excel or CSV File", type=["xlsx", "xls", "csv"]
)

office_list_path = "OfficeList.xlsx"

try:
  office_df = pd.read_excel(office_list_path)
except Exception as e:
  st.error(f"Error loading OfficeList.xlsx from repository: {e}")
  office_df = None

if uploaded_file is not None and office_df is not None:
  try:
    if uploaded_file.name.endswith(".csv"):
      daily_df = pd.read_csv(uploaded_file)
    else:
      daily_df = pd.read_excel(uploaded_file)

    st.success("Files loaded successfully! Processing report...")

    # --- 1. EXTRACT TARGET OFFICES FROM OFFICE LIST ---
    target_offices = set(
        office_df.iloc[:, 0].dropna().astype(str).str.strip().str.lower()
    )
    office_name_mapping = {
        name.lower(): name
        for name in office_df.iloc[:, 0].dropna().astype(str).str.strip()
    }

    # --- 2. MAP COLUMNS BASED ON VBA SPECIFICATION ---
    processed_rows = []

    for idx, row in daily_df.iterrows():
      raw_name = str(row.iloc[1]).strip() if len(row) > 1 else ""
      lookup_name = raw_name.lower()

      if lookup_name in target_offices:
        official_name = office_name_mapping[lookup_name]

        cash_val = float(row.iloc[3]) if len(row) > 3 and pd.notnull(row.iloc[3]) else 0.0
        dqr_val = float(row.iloc[15]) if len(row) > 15 and pd.notnull(row.iloc[15]) else 0.0
        pos_card_val = float(row.iloc[19]) if len(row) > 19 and pd.notnull(row.iloc[19]) else 0.0
        pos_qr_val = float(row.iloc[21]) if len(row) > 21 and pd.notnull(row.iloc[21]) else 0.0
        epay_val = float(row.iloc[25]) if len(row) > 25 and pd.notnull(row.iloc[25]) else 0.0

        total_digi = dqr_val + pos_card_val + pos_qr_val + epay_val
        total_trans = cash_val + total_digi

        if total_trans > 0:
          pct_digi = (total_digi / total_trans) if total_trans > 0 else 0.0
          processed_rows.append({
              "Office Name": official_name,
              "Cash (Cnt)": int(cash_val),
              "DQR Scan (Cnt)": int(dqr_val),
              "SBI POS-CARD (Cnt)": int(pos_card_val),
              "SBI POS BHARAT QR (Cnt)": int(pos_qr_val),
              "SBI E PAY UPI (Cnt)": int(epay_val),
              "Total Digital Transactions": int(total_digi),
              "Total Transactions": int(total_trans),
              "% of Digital Transactions": pct_digi,
          })

    result_df = pd.DataFrame(processed_rows)

    if result_df.empty:
      st.warning("No matching records found with transactions greater than 0!")
    else:
      # --- 3. THREE-LEVEL SORTING ---
      result_df = result_df.sort_values(
          by=[
              "% of Digital Transactions",
              "Cash (Cnt)",
              "Total Digital Transactions",
          ],
          ascending=[True, False, True],
      ).reset_index(drop=True)

      # --- 4. BUILD EXCEL WORKBOOK ---
      output = io.BytesIO()
      wb = openpyxl.Workbook()
      ws = wb.active
      ws.title = "DailyReport"

      report_date = (datetime.now() - timedelta(days=1)).strftime("%d-%m-%Y")

      # Title Banner (A1:I1)
      ws.merge_cells("A1:I1")
      title_cell = ws["A1"]
      title_cell.value = (
          f"Nellore Division: Digital Transactions information dated"
          f" {report_date}"
      )
      title_cell.font = Font(
          name="Calibri", size=14, bold=True, color="FFFFFF"
      )
      title_cell.fill = PatternFill(
          start_color="0F2043", end_color="0F2043", fill_type="solid"
      )
      title_cell.alignment = Alignment(
          horizontal="center", vertical="center"
      )
      ws.row_dimensions[1].height = 30

      # Table Headers (Row 3 & 4)
      headers = [
          "Office Name",
          "Cash (Cnt)",
          "DQR Scan (Cnt)",
          "SBI POS-CARD (Cnt)",
          "SBI POS BHARAT QR (Cnt)",
          "SBI E PAY UPI (Cnt)",
          "Total Digital Transactions",
          "Total Transactions",
          "% of Digital Transactions",
      ]
      sub_headers = [
          "a",
          "b",
          "c",
          "d",
          "e",
          "f",
          "g=c+d+e+f",
          "h=b+g",
          "i=(g/h)*100",
      ]

      for col_num, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_num, value=h)
        cell.font = Font(name="Calibri", size=11, bold=True)
        cell.fill = PatternFill(
            start_color="A9D08E", end_color="A9D08E", fill_type="solid"
        )
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )

      for col_num, sh in enumerate(sub_headers, 1):
        cell = ws.cell(row=4, column=col_num, value=sh)
        cell.font = Font(name="Calibri", size=11, bold=True)
        cell.fill = PatternFill(
            start_color="A9D08E", end_color="A9D08E", fill_type="solid"
        )
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )

      # Row 3 height updated to 63 as requested
      ws.row_dimensions[3].height = 63
      ws.row_dimensions[4].height = 20

      # Insert Data Rows
      start_row = 5
      for idx, row_data in result_df.iterrows():
        current_row = start_row + idx
        ws.cell(row=current_row, column=1, value=row_data["Office Name"])
        ws.cell(
            row=current_row, column=2, value=row_data["Cash (Cnt)"]
        ).number_format = "#,##0"
        ws.cell(
            row=current_row, column=3, value=row_data["DQR Scan (Cnt)"]
        ).number_format = "#,##0"
        ws.cell(
            row=current_row, column=4, value=row_data["SBI POS-CARD (Cnt)"]
        ).number_format = "#,##0"
        ws.cell(
            row=current_row, column=5, value=row_data["SBI POS BHARAT QR (Cnt)"]
        ).number_format = "#,##0"
        ws.cell(
            row=current_row, column=6, value=row_data["SBI E PAY UPI (Cnt)"]
        ).number_format = "#,##0"

        ws.cell(
            row=current_row,
            column=7,
            value=f"=SUM(C{current_row}:F{current_row})",
        ).number_format = "#,##0"
        ws.cell(
            row=current_row, column=8, value=f"=B{current_row}+G{current_row}"
        ).number_format = "#,##0"
        ws.cell(
            row=current_row,
            column=9,
            value=f"=IFERROR(G{current_row}/H{current_row}, 0)",
        ).number_format = "0.00%"

        pct = row_data["% of Digital Transactions"]
        fill_color = "FA696B"
        if pct >= 0.75:
          fill_color = "8AD090"
        elif pct >= 0.50:
          fill_color = "FFEBF4"
        elif pct >= 0.25:
          fill_color = "FCAAA6"

        ws.cell(row=current_row, column=9).fill = PatternFill(
            start_color=fill_color, end_color=fill_color, fill_type="solid"
        )

      end_row = start_row + len(result_df) - 1

      # Total Row
      total_row = end_row + 1
      ws.cell(row=total_row, column=1, value="Total")
      for c_idx in range(2, 9):
        col_letter = openpyxl.utils.get_column_letter(c_idx)
        ws.cell(
            row=total_row,
            column=c_idx,
            value=f"=SUM({col_letter}{start_row}:{col_letter}{end_row})",
        ).number_format = "#,##0"

      ws.cell(
          row=total_row,
          column=9,
          value=f"=IFERROR(G{total_row}/H{total_row}, 0)",
      ).number_format = "0.00%"

      # Formatting borders, fonts & alignment
      thin_border = Border(
          left=Side(style="thin", color="000000"),
          right=Side(style="thin", color="000000"),
          top=Side(style="thin", color="000000"),
          bottom=Side(style="thin", color="000000"),
      )

      for r in range(3, total_row + 1):
        for c in range(1, 10):
          cell = ws.cell(row=r, column=c)
          cell.font = Font(name="Calibri", size=11)
          cell.border = thin_border
          if c > 1:
            cell.alignment = Alignment(horizontal="center", vertical="center")
          else:
            cell.alignment = Alignment(horizontal="left", vertical="center")

      for c in range(1, 10):
        t_cell = ws.cell(row=total_row, column=c)
        t_cell.font = Font(name="Calibri", size=11, bold=True)
        t_cell.fill = PatternFill(
            start_color="E6E6E6", end_color="E6E6E6", fill_type="solid"
        )
        t_cell.border = Border(
            top=Side(style="thin", color="000000"),
            bottom=Side(style="double", color="000000"),
            left=Side(style="thin", color="000000"),
            right=Side(style="thin", color="000000"),
        )

      ws.column_dimensions["A"].width = 30
      for col_char in ["B", "C", "D", "E", "F", "G", "H", "I"]:
        ws.column_dimensions[col_char].width = 14

      wb.save(output)
      processed_data = output.getvalue()

      st.success(
          f"Successfully processed {len(result_df)} active offices! Zero-transaction"
          " offices eliminated."
      )

      st.write("Preview of Formatted Report:", result_df.head())

      st.download_button(
          label="📥 Download Formatted Daily Report",
          data=processed_data,
          file_name=(
              "63_Offices_Digital_Report_"
              f"{(datetime.now() - timedelta(days=1)).strftime('%Y%m%d')}.xlsx"
          ),
          mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
      )

  except Exception as e:
    st.error(f"Error processing the uploaded file: {e}")
